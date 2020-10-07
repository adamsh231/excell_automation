# %%
import openpyxl
from openpyxl import Workbook
import csv
import os


class Data:
    def __init__(self, fileName, sheetName):
        self.fileName = fileName
        self.sheetName = sheetName

    def setData(self, data):
        self.data = data


# %%
data = list()
path = "E:\\TUGAS\\PROJECT BESAR\\Kode\\Python\\filecsv"

# %%
for csvFileName in os.listdir(path):
    if not csvFileName.endswith('.csv'):
        continue
    customizeSheetName = (csvFileName.replace('.csv', '')).split('-')[0]
    data.append(Data(csvFileName, customizeSheetName))
for i in range(len(data)):
    openFile = open(path+"\\"+data[i].fileName)
    readFile = csv.reader(openFile)
    fileData = list(readFile)
    data[i].setData(fileData)
    openFile.close()

# %%
workBook = Workbook()
workSheet = workBook.active
workSheet.title = "131098"
for i in range(len(data)):
    workBook.create_sheet(title=data[i].sheetName)

workBook.save(filename=path+"\\Template.xlsx")

# %%
workBook = openpyxl.load_workbook(filename=path+"\\Template.xlsx")

for i in range(len(data)):
    sheet = workBook.get_sheet_by_name(data[i].sheetName)
    for j in range(len(data[i].data)):
        for k in range(len(data[i].data[0])):
            if (k == 1 or k == 4) and j > 0:
                data[i].data[j][k] = float(data[i].data[j][k])
            sheet.cell(row=(j+1),column=(k+1)).value = data[i].data[j][k]
workBook.save(filename=path+"\\Template.xlsx")

#%%
